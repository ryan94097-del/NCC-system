# -*- coding: utf-8 -*-
"""產生去識別化測試樣本 Excel，重現真實 RCB 登錄表結構與資料坑。"""
import os
from openpyxl import Workbook

OUT_DIR = r"c:\Users\Ryan\Documents\Antigravity folder\NCC surveillance\sample"
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "NCC_測試樣本.xlsx")

HEADERS = ['委託編號','委託日期','委託機構','委託產品','廠牌','型號','委託項目','繳費單',
           '接件人員','委派人員','驗證人員\n接收日期','補件通知\n通知時間/連絡人員','結果',
           '證書編號','證書日期','繳款代號','發文','開立發票','NCC登錄','聯絡人電子郵件',
           '繳款單繳費者','備註','繳款單已情款\n(PA # )','是否跟NCC請款','收到原稿TRT合約書',
           'TN證書號碼','LH No.','CD燒錄號碼']
NCOL = len(HEADERS)

def row(**kw):
    """依欄位名填一列，其餘留空。"""
    r = [None]*NCOL
    for k,v in kw.items():
        r[HEADERS.index(k)] = v
    return r

def make_sheet(ws, big_title, data_rows):
    ws.append([big_title] + [None]*(NCOL-1))   # row0 大標題
    ws.append(HEADERS)                          # row1 真正表頭
    for r in data_rows:
        ws.append(r)

wb = Workbook()

# ---- LPD 分頁 ----
ws_lpd = wb.active
ws_lpd.title = "LPD型式認證列表"
lpd_rows = [
    # 2026 有效案件（去識別化假資料）
    row(委託編號='114100001', 委託日期='2026.01.10', 委託機構='艾克美股份有限公司', 委託產品='藍牙耳機',
        廠牌='ACME', 型號='AX-100', 證書編號='CCAN26LP0010T1', 接件人員='Tester A'),
    row(委託編號='114100002', 委託日期='2026.01.12', 委託機構='福科技股份有限公司', 委託產品='無線滑鼠',
        廠牌='FooTech', 型號='FT-2200', 證書編號='CCAN26LP0020T4', 接件人員='Tester A'),
    row(委託編號='114100003', 委託日期='2026.02.03', 委託機構='貝爾科技', 委託產品='智慧溫濕度感應器',
        廠牌='BrandX', 型號='溫濕度感應器 TH-9', 證書編號='CCAN26LP0030T7', 接件人員='Tester C'),
    # 通用短型號（測試關鍵字補強）
    row(委託編號='114100004', 委託日期='2026.03.01', 委託機構='通用電子', 委託產品='車用多媒體盒',
        廠牌='GenCorp', 型號='X8', 證書編號='CCAN26LP0040T0', 接件人員='Tester C'),
    # 換行證號（測試清洗）
    row(委託編號='114100005', 委託日期='2026.03.15', 委託機構='通用電子', 委託產品='車用多媒體盒',
        廠牌='GenCorp', 型號='X9', 證書編號='CCAN26LP0041T9\n(系列1)', 接件人員='Tester C'),
    # 2025 案件（測試年份過濾：應被濾掉）
    row(委託編號='114090001', 委託日期='2025.05.20', 委託機構='宜家風', 委託產品='溫度濕度感應器',
        廠牌='IKEAlike', 型號='TIMMER-01', 證書編號='CCAN25LP0010T6', 接件人員='Tester A'),
    row(委託編號='114090002', 委託日期='2025.06.20', 委託機構='矽力', 委託產品='嵌入式無線模組',
        廠牌='SilexLike', 型號='SX-USBAC', 證書編號='CCAN25Y10020T0', 接件人員='Tester A'),
    # 年份分隔列（整列只有一個數字，應丟棄）
    row(委託編號='2018'),
    # NCC 請款彙總列（證書編號為空，應丟棄）
    row(委託機構='電信終端設備審查費用(115年02月)', 委託項目='NCC 請款'),
    # 空型號（應丟棄）
    row(委託編號='114100099', 委託日期='2026.04.01', 委託機構='空型號測試', 委託產品='未知裝置',
        廠牌='NoModel', 型號=None, 證書編號='CCAN26LP0099T2'),
]
make_sheet(ws_lpd, 'RCB工服委託登錄表 (LPD)', lpd_rows)

# ---- TTE 分頁 ----
ws_tte = wb.create_sheet("TTE型式認證列表")
tte_rows = [
    row(委託編號='114200001', 委託日期='2026.01.20', 委託機構='寵物科技', 委託產品='寵物穿戴式定位通訊器',
        廠牌='PetGlobal', 型號='PG-24A01', 證書編號='CCAN264G0010T8', 接件人員='Tester E'),
    row(委託編號='114200002', 委託日期='2026.02.10', 委託機構='車聯', 委託產品='車用多媒體盒',
        廠牌='AutoLink', 型號='ALX-8', 證書編號='CCAN264G0020T1', 接件人員='Tester E'),
    # 換行證號 + 系列
    row(委託編號='114200003', 委託日期='2026.02.11', 委託機構='車聯', 委託產品='車用多媒體盒',
        廠牌='AutoLink', 型號='ALX-9', 證書編號='CCAN264G0021T0\n(系列1)', 接件人員='Tester E'),
    # 年份分隔列
    row(委託編號='2018'),
    # 2025（應被年份過濾）
    row(委託編號='114190001', 委託日期='2025.03.10', 委託機構='連結', 委託產品='4G LTE 無線路由器',
        廠牌='DLinkLike', 型號='G403C', 證書編號='CCAN254G0010T0', 接件人員='Tester E'),
]
make_sheet(ws_tte, 'RCB工服委託登錄表 (TTE)', tte_rows)

# ---- 干擾分頁（不應被處理）----
ws_other = wb.create_sheet("TRT Contract Date")
ws_other.append(['TRT Contract Date', None, None, None])
ws_other.append(['委託機構', 'English', 'LH No.', 'Contract date'])
ws_other.append(['艾克美股份有限公司', None, None, None])

wb.save(OUT)
print("WROTE:", OUT)

# 驗證讀回
import pandas as pd
xl = pd.ExcelFile(OUT, engine='openpyxl')
print("SHEETS:", xl.sheet_names)
for sn in ['LPD型式認證列表','TTE型式認證列表']:
    df = pd.read_excel(xl, sheet_name=sn, header=1, dtype=str)
    print(f"  {sn}: {len(df)} rows, 有證號 {df['證書編號'].notna().sum()}")
